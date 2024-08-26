import sys
import os
import requests
from PyQt5.QtWidgets import (QApplication, QWidget, QVBoxLayout, QHBoxLayout, QLineEdit,
                             QPushButton, QTableWidget, QTableWidgetItem, QLabel, QMessageBox,
                             QHeaderView, QFrame)
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QPixmap, QFont, QColor, QPalette
from bs4 import BeautifulSoup
from io import BytesIO
import re
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter


class ChurchInfoApp(QWidget):
    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        # 전체 레이아웃
        main_layout = QHBoxLayout()

        # 왼쪽 프레임 (6)
        left_frame = QFrame()
        left_layout = QVBoxLayout()

        # 구분 입력 필드
        self.category_input = QLineEdit(self)
        self.category_input.setPlaceholderText('구분')
        self.category_input.setFont(QFont('Arial', 10, QFont.Bold))

        # URL 입력 필드
        self.url_input = QLineEdit(self)
        self.url_input.setPlaceholderText('URL을 입력하세요...')
        self.url_input.setFont(QFont('Arial', 10, QFont.Bold))
        self.url_input.returnPressed.connect(self.fetch_info)

        # 버튼 레이아웃
        button_layout = QHBoxLayout()

        # 정보 가져오기 버튼
        self.fetch_button = QPushButton('정보 가져오기', self)
        self.fetch_button.setFont(QFont('Arial', 10, QFont.Bold))
        self.fetch_button.clicked.connect(self.fetch_info)

        # 초기화 버튼
        self.clear_button = QPushButton('초기화', self)
        self.clear_button.setFont(QFont('Arial', 10, QFont.Bold))
        self.clear_button.clicked.connect(self.clear_fields)

        # 버튼 레이아웃에 버튼 추가
        button_layout.addWidget(self.fetch_button)
        button_layout.addWidget(self.clear_button)

        # 왼쪽 레이아웃에 요소 추가
        left_layout.addWidget(QLabel('구분:'))
        left_layout.addWidget(self.category_input)
        left_layout.addWidget(QLabel('URL:'))
        left_layout.addWidget(self.url_input)
        left_layout.addLayout(button_layout)
        left_frame.setLayout(left_layout)

        # 오른쪽 프레임 (4)
        right_frame = QFrame()
        right_layout = QVBoxLayout()

        # 로고 및 타이틀 배치
        logo_label = QLabel(self)
        pixmap = QPixmap('logo.png')
        if not pixmap.isNull():
            logo_label.setPixmap(pixmap.scaled(100, 100, Qt.KeepAspectRatio))
        right_layout.addWidget(logo_label, alignment=Qt.AlignCenter)

        title_label = QLabel('남경기노회', self)
        title_label.setFont(QFont('Arial', 20))
        right_layout.addWidget(title_label, alignment=Qt.AlignCenter)
        right_frame.setLayout(right_layout)

        # 메인 레이아웃에 프레임 추가
        main_layout.addWidget(left_frame, 6)
        main_layout.addWidget(right_frame, 4)

        # 테이블 설정
        self.table = QTableWidget(self)
        self.table.setColumnCount(7)
        self.table.setHorizontalHeaderLabels(
            ['사진', '이름', '교회명', '우편번호', '주소', '전화번호', '이메일'])

        # 헤더 색상 설정
        header = self.table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.Stretch)
        palette = self.table.palette()
        palette.setColor(QPalette.Base, QColor("#c7c8c9"))
        self.table.setPalette(palette)
        self.table.horizontalHeader().setStyleSheet(
            "QHeaderView::section { background-color: #c7c8c9; }")

        # 전체 레이아웃 구성
        layout = QVBoxLayout()
        layout.addLayout(main_layout)
        layout.addWidget(self.table)

        # 인용구 섹션 추가
        footer_label = QLabel('made by 나종춘(2024)', self)
        footer_label.setFont(QFont('Arial', 10))
        footer_label.setAlignment(Qt.AlignRight)
        layout.addWidget(footer_label)

        self.setLayout(layout)
        self.setWindowTitle('교회 정보 추출기')
        self.setGeometry(300, 300, 800, 600)
        self.center()

    def center(self):
        # 화면의 가운데에 창을 배치
        qr = self.frameGeometry()
        cp = QApplication.desktop().screenGeometry().center()
        qr.moveCenter(cp)
        self.move(qr.topLeft())

    def fetch_info(self):
        category = self.category_input.text().strip()
        url = self.url_input.text().strip()
        if not url:
            QMessageBox.warning(self, '입력 오류', 'URL을 입력하세요.')
            return

        if not url.startswith("http://") and not url.startswith("https://"):
            url = "http://" + url

        try:
            response = requests.get(url)
            response.raise_for_status()  # HTTP 오류가 있는지 확인
        except requests.exceptions.RequestException as e:
            QMessageBox.critical(
                self, 'URL 오류', f'URL을 가져오는 중 오류가 발생했습니다: {e}')
            return

        # HTML 파싱 및 데이터 추출
        soup = BeautifulSoup(response.text, 'html.parser')

        blurbs = soup.find_all('div', class_='et_pb_blurb_content')
        data_list = []

        # '구분' 데이터 추가
        if category:
            category_item = QTableWidgetItem(f"--{category}--")
            category_item.setForeground(QColor('blue'))
            category_item.setTextAlignment(Qt.AlignCenter)
            category_item.setFont(QFont('Arial', 12, QFont.Bold))
            data_list.append(
                [category_item, None, None, None, None, None, None])

        for blurb in blurbs:
            img_tag = blurb.find('img')
            name_tag = blurb.find('h4', class_='et_pb_module_header')
            blurb_description = blurb.find(
                'div', class_='et_pb_blurb_description')
            if blurb_description:
                church_name_tag = blurb_description.find(
                    string=lambda t: '교회' in t)
                address_tags = blurb_description.find_all('br')
                tel_tags = blurb_description.find_all(
                    'a', href=lambda h: 'tel' in h)
                email_tag = blurb_description.find(
                    'a', href=lambda h: 'mailto' in h)

                # 필터링 (교회명이 있는 경우만)
                if church_name_tag and '교회명칭 표기 오류' not in church_name_tag:
                    img = img_tag['src'] if img_tag else ''
                    # 이름에서 숫자와 점 제거
                    name = re.sub(
                        r'^\d+\.', '', name_tag.text).strip() if name_tag else ''
                    church_name = church_name_tag.strip()
                    postcode = re.search(r'\d{5}', blurb_description.get_text()).group(
                        0) if re.search(r'\d{5}', blurb_description.get_text()) else ''
                    address = re.search(r'\d{5}\s*(.*)', blurb_description.get_text()).group(
                        1).strip() if re.search(r'\d{5}\s*(.*)', blurb_description.get_text()) else ''
                    tel = ', '.join(
                        [tel.text for tel in tel_tags]) if tel_tags else ''
                    email = email_tag.text if email_tag else ''

                    data_list.append(
                        [img, name, church_name, postcode, address, tel, email])

        # 테이블 업데이트
        row_position = self.table.rowCount()
        self.table.setRowCount(row_position + len(data_list))

        for row_num, data in enumerate(data_list):
            if isinstance(data[0], QTableWidgetItem):  # '구분' 데이터인 경우
                self.table.setItem(row_position + row_num, 0, data[0])
            else:
                for col_num, item in enumerate(data):
                    if col_num == 0 and item:  # 사진 셀의 경우
                        img_data = requests.get(item).content
                        pixmap = QPixmap()
                        pixmap.loadFromData(BytesIO(img_data).read())
                        if not pixmap.isNull():
                            img_label = QLabel()
                            img_label.setPixmap(pixmap.scaled(
                                100, 100, Qt.KeepAspectRatio))
                            img_label.setAlignment(Qt.AlignCenter)
                            self.table.setCellWidget(
                                row_position + row_num, col_num, img_label)
                    else:
                        cell = QTableWidgetItem(item)
                        if col_num == 1 or col_num == 3:  # 이름과 우편번호를 중앙 정렬
                            cell.setTextAlignment(Qt.AlignCenter)
                        self.table.setItem(
                            row_position + row_num, col_num, cell)

        # 데이터를 엑셀 파일로 저장
        self.save_to_excel(data_list)

    def save_to_excel(self, data_list):
        # 엑셀 파일 생성 및 열기
        file_path = "members_list.xlsx"
        if os.path.exists(file_path):
            wb = load_workbook(file_path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Members"
            ws.append(['사진', '이름', '교회명', '우편번호',
                      '주소', '전화번호', '이메일'])  # 헤더 추가

        # 마지막 빈 행을 찾음
        row_num = ws.max_row + 1

        for data in data_list:
            if isinstance(data[0], QTableWidgetItem):  # '구분' 데이터인 경우
                row_data = [data[0].text()] + [''] * 6  # 나머지 셀 빈칸 채우기
                ws.append(row_data)
                row_num += 1
            else:
                for col_num, item in enumerate(data):
                    if col_num == 0 and item:  # 사진 데이터
                        img = Image(BytesIO(requests.get(item).content))
                        img.width = 80
                        img.height = 80
                        cell_ref = f'A{row_num}'
                        ws.add_image(img, cell_ref)
                        ws.row_dimensions[row_num].height = 80  # 행 높이 조정
                    else:
                        cell_ref = f'{get_column_letter(col_num + 1)}{row_num}'
                        ws[cell_ref] = item
                row_num += 1

        # 열 너비 조정
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        # 엑셀 파일 저장
        wb.save(file_path)
        QMessageBox.information(self, '저장 완료', f'데이터가 {
                                file_path} 파일에 저장되었습니다.')

    def clear_fields(self):
        # 입력 필드와 테이블 초기화
        self.category_input.clear()
        self.url_input.clear()
        self.table.setRowCount(0)


if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = ChurchInfoApp()
    ex.show()
    sys.exit(app.exec_())
